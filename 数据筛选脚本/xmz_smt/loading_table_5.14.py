import xlrd
import pandas as pd
import numpy as np
import glob
import os
import re
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def process_merged_cells_xls(file_path, sheet_name):
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_name(sheet_name)
    
    merged_regions = []
    for merged_range in sheet.merged_cells:
        min_row, max_row, min_col, max_col = merged_range
        value = sheet.cell(min_row, min_col).value
        merged_regions.append((min_row, max_row, min_col, max_col, value))
    
    data = []
    for row_idx in range(sheet.nrows):
        row_data = []
        for col_idx in range(sheet.ncols):
            cell_value = sheet.cell(row_idx, col_idx).value
            for (m_min_row, m_max_row, m_min_col, m_max_col, m_value) in merged_regions:
                if (m_min_row <= row_idx < m_max_row) and (m_min_col <= col_idx < m_max_col):
                    cell_value = m_value
                    break
            row_data.append(cell_value)
        data.append(row_data)
    return data

def extract_data_blocks(data):
    blocks = []
    current_block = None
    direction_search_range = 3  # 在标题行后最多搜索3行寻找方向

    for i, row in enumerate(data):
        # 匹配包含 Line 的标题行
        line_cells = [cell for cell in row if isinstance(cell, str) and 'Line' in cell]

        if line_cells:
            # 保存前一个区块
            if current_block:
                blocks.append(current_block)
            
            # 初始化新区块
            title_cell = line_cells[0]
            current_block = {
                'title': title_cell,
                'direction': None,  # 初始化为空
                'headers': [],
                'rows': [],
                'direction_found': False  # 方向是否已找到的标记
            }

            # 在后续行中搜索方向信息（最多检查3行）
            for j in range(i+1, min(i+1 + direction_search_range, len(data))):
                direction_row = data[j]
                # 匹配独立的方向单元格（纯"左"或"右"）
                direction_cells = [
                    cell for cell in direction_row 
                    if isinstance(cell, str) and re.fullmatch(r'\s*[左右]\s*', str(cell))
                ]
                if direction_cells:
                    current_block['direction'] = direction_cells[0].strip()
                    current_block['direction_found'] = True
                    break  # 找到后停止搜索

        elif current_block:
            # 优先处理方向未找到的情况
            if not current_block['direction_found']:
                # 尝试在本行匹配复合方向信息（如"方向：左"）
                for cell in row:
                    if isinstance(cell, str):
                        match = re.search(r'([左右])', cell)
                        if match:
                            current_block['direction'] = match.group(1)
                            current_block['direction_found'] = True
                            break

            # 列头识别逻辑（仅在方向处理后执行）
            if not current_block['headers']:
                if any('供料器' in str(cell) for cell in row):
                    current_block['headers'] = [str(cell).strip() for cell in row]
                    continue  # 跳过本行后续处理

            # 数据行收集（排除空行）
            if any(cell is not None for cell in row):
                current_block['rows'].append(row)

    # 添加最后一个区块
    if current_block:
        blocks.append(current_block)
    
    return blocks

def daochu():
# 主流程
    bom_folder = os.path.join(os.getcwd(), '导出文件')
    files = glob.glob(os.path.join(bom_folder, '*.xls'))
    if not files:
        print("未找到 .xls 文件，请检查路径！")
        exit()

    file_path = files[0]
    data = process_merged_cells_xls(file_path, 'Sheet1')

    # 调试数据
    # print("==== 前10行数据 ====")
    # for idx, row in enumerate(data[:10]):
    #     print(f"行 {idx}: {row}")

    data_blocks = extract_data_blocks(data)
    print("提取到的数据块数量:", len(data_blocks))

    # 转换为 DataFrame
    dfs = []
    for block in data_blocks:
        if block.get('headers') and block.get('rows'):
            df = pd.DataFrame(block['rows'], columns=block['headers'])
            df['料台标题'] = block['title']
            df['料台左右'] = block['direction']
            dfs.append(df)
        # else:
        #     print("跳过无效块:", block)

    if dfs:
        final_df = pd.concat(dfs, ignore_index=True)
        
        # --------------------------
        # 新增输出结果处理逻辑
        # --------------------------
        # 1. B列（供料器列）向下填充空值
        if '供料器' in final_df.columns:
            final_df['供料器'] = final_df['供料器'].fillna(method='ffill')
        else:
            print("警告: 未找到B列（供料器列）")
        
        target_column = '元件'  # 根据实际列名修改

        # 2. 空值预处理（处理多种空值表现形式）
        null_values = ['', 'NA', 'NaN', 'NULL', 'None', ' ']
        final_df[target_column] = final_df[target_column].replace(null_values, pd.NA)

        # 3. 统计空值情况（调试用）
        # print(f"删除前数据量: {len(final_df)} 行")
        # print(f"'{target_column}'列空值数量: {final_df[target_column].isna().sum()}")

        # 4. 执行删除操作（保留有数据的行）
        if target_column in final_df.columns:
            # 方法一：直接删除空值行（严格模式）
            final_df.dropna(
                subset=[target_column],
                how='any',
                inplace=True
            )
            
            
            # print(f"删除后数据量: {len(final_df)} 行")
            
            # 二次验证
            if final_df[target_column].isna().sum() > 0:
                print("警告: 仍存在空值，请检查数据源")
        else:
            print(f"关键列 '{target_column}' 不存在，可用列：{list(final_df.columns)}")

        # 自动填充空白列名（优化版）
        new_columns = []
        counter = 0
        for col in final_df.columns:
            # 处理空值或默认生成的 Unnamed 列
            if pd.isna(col) or str(col).strip() in ('', 'Unnamed: 0'):
                new_col = f"column_{counter}"
                counter += 1
            else:
                new_col = col
            
            # 统一中文列名（示例：包含 "Line" 的列重命名为 "料台标题"）
            if 'Line' in str(new_col):
                new_col = '料台标题'
            
            new_columns.append(new_col)

        final_df.columns = new_columns

        # 验证列名是否存在（若缺失则动态创建空列）
        required_columns = ["料台标题", "料台左右", "料槽", "供料器类型", "column_8", "元件", "间距"]
        for col in required_columns:
            if col not in final_df.columns:
                final_df[col] = None  # 创建空列或填充默认值

        # 正确选择多列（使用列表语法）
        end_df = final_df[required_columns]

        end_df.replace(['', 'NA', 'NaN', 'null', 'None', ' '], np.nan, inplace=True)

        # 检查数据类型并转换
        for col in end_df.columns:
            if end_df[col].dtype == 'category':
                end_df[col] = end_df[col].astype('object')

        # 重置索引
        end_df = end_df.reset_index(drop=True)

        # 填充NaN，先ffill再bfill处理首行
        filleed_df = end_df.fillna(method='ffill').fillna(method='bfill')

        
        def merge_flag(row):
            name = row['料台标题']
            left_right = row['料台左右']
            if left_right == '左':
                return f"{name}-2"
            elif left_right == '右':
                return f"{name}-1"
        filleed_df['sheet表名'] = filleed_df.apply(merge_flag, axis=1)

        def merge_with_flag(row):
            name = row['sheet表名']
            cao = row['料槽']
            price = row['column_8']
            return f"{name}-{int(cao)}-{int(price)}"[6:]

        filleed_df['合并结果'] = filleed_df.apply(merge_with_flag, axis=1)
        
        # 保存结果
        filleed_df.to_excel("清洗后的导出文件.xlsx",index=False)
        # print("处理完成，已保存到 output.csv")
        return filleed_df
    else:
        print("警告: 无有效数据块可合并！")













def bom():

    logging.basicConfig(level=logging.INFO)

    # 配置路径和关键词
    bom_folder = os.path.join(os.getcwd(), 'BOM')
    files = glob.glob(os.path.join(bom_folder, '*[Bb][Oo][Mm]*.xlsx'))
    required_columns = {'编码', '描述', '位号'}
    keyword_map = {'编码': '物料编码', '描述': '物料描述', '位号': '位号', '用量': '用量'}

    data_frames = []

    for file in files:
        try:
            xls = pd.ExcelFile(file)
        except Exception as e:
            logging.error(f"读取文件失败: {file}, 错误: {e}")
            continue
        
        for sheet_name in xls.sheet_names:
            try:
                # 动态寻找表头行（前10行内）
                header_row = None
                for skiprows in range(0, 10):
                    df_temp = pd.read_excel(xls, sheet_name, skiprows=skiprows, nrows=1)
                    row_columns = df_temp.columns.str.lower().tolist()
                    # 检查是否包含所有必需列的关键词
                    found = all(any(kw in col for col in row_columns) for kw in required_columns)
                    if found:
                        header_row = skiprows
                        break
                
                if header_row is None:
                    logging.info(f"跳过 {file} 的 {sheet_name}: 未找到表头")
                    continue
                
                # 读取数据
                df = pd.read_excel(xls, sheet_name, skiprows=header_row)
                if df.empty:
                    logging.info(f"文件 {file} 的 {sheet_name} 无数据")
                    continue
                
                # 添加来源信息
                df['来源文件'] = os.path.basename(file)
                df['来源Sheet'] = sheet_name
                # --- 修改后的合并位号列代码 ---
                ref_cols = [col for col in df.columns if '位号' in col]  # 改为包含"位号"的列名
                if len(ref_cols) > 0:
                    def merge_ref_des(row):
                        refs = []
                        for col in ref_cols:
                            val = row[col]
                            if pd.notna(val):
                                # 处理可能的换行符和多个空格
                                parts = str(val).replace('\n', ' ').split()  # 自动处理连续空格和换行
                                refs.extend([p.strip() for p in parts if p.strip()])
                        # 按原始顺序去重（替代set的无序性）
                        seen = set()
                        ordered_refs = []
                        for ref in refs:
                            if ref not in seen:
                                seen.add(ref)
                                ordered_refs.append(ref)
                        return ' '.join(ordered_refs) if ordered_refs else ''
                    
                    df['位号'] = df.apply(merge_ref_des, axis=1)
                    df.drop(ref_cols, axis=1, inplace=True)
                data_frames.append(df)
                    
                # --- 列名重命名（确保唯一性）---
                sorted_keywords = sorted(keyword_map.keys(), key=lambda x: -len(x))
                new_columns = []
                seen = set()
                for col in df.columns:
                    col_str = str(col)
                    col_lower = col_str.lower()
                    matched = False
                    for kw in sorted_keywords:
                        if kw in col_lower:
                            base_name = keyword_map[kw]
                            if base_name in seen:
                                suffix = 1
                                while f"{base_name}.{suffix}" in seen:
                                    suffix += 1
                                new_name = f"{base_name}.{suffix}"
                            else:
                                new_name = base_name
                            new_columns.append(new_name)
                            seen.add(new_name)
                            matched = True
                            break
                    if not matched:
                        new_columns.append(col_str)
                        seen.add(col_str)
                df.columns = new_columns
            except Exception as e:
                logging.warning(f"处理 {file} 的 {sheet_name} 出错: {e}")

    # 合并和清洗数据
    if data_frames:
        combined_df = pd.concat(data_frames, ignore_index=True)
        combined_df = combined_df.dropna(axis=1, how='all')
        final_columns = list(keyword_map.values()) + ['来源文件', '来源Sheet']
        final_df = combined_df.reindex(columns=final_columns, fill_value='')
        
        final_df['位号'] = (
        final_df['位号']
        .astype(str)
        .replace(['nan', '', ' '], pd.NA)  # 统一替换为 Pandas 空值
    )
        final_df['位号'] = final_df['位号'].fillna(method='ffill')
        # print("合并后的数据样例：")
        # print(final_df.head())
        final_df['替代'] = final_df['物料编码'].where(final_df['用量'].notna(), other=pd.NA)
        final_df['替代'] = final_df['替代'].fillna(method='ffill')
        final_df.to_excel("bom.xlsx",index=False)
        return final_df
    else:
        print("未找到有效数据")


daochu_c = daochu()
bom_c = bom()


merged_df = pd.merge(
    bom_c,                     # 左表
    daochu_c,                  # 右表
    left_on="替代",         # 左表连接键列名
    right_on="元件",            # 右表连接键列名
    how="outer",               # 外连接
    suffixes=('_bom', '_daochu')  # 列名后缀（可选，避免重复列名冲突）
)

merged_df.to_excel("合并后数据.xlsx")
# 1. 保留“料台标题”不为空的数据（假设该列存在于合并后的表中）
filtered_df = merged_df.dropna(subset=['料台标题','物料编码'])



# 清理Sheet名称的非法字符（Excel限制）
def clean_sheet_name(name):
    # 替换非法字符为空格，并截断长度
    cleaned = re.sub(r'[\\/?:*$$$$]', '', str(name)).strip()  # 去除非字母、数字、下划线的字符
    return cleaned[:31]  # Excel sheet名称最长31字符

# 获取分类后的唯一Sheet表名（并处理重复）
unique_sheets = filtered_df["sheet表名"].unique()
sheet_names = []
seen = {}  # 记录已使用的名称

for sheet in unique_sheets:
    clean_name = clean_sheet_name(sheet)
    if clean_name in seen:
        seen[clean_name] += 1
        final_name = f"{clean_name}_{seen[clean_name]}"
    else:
        seen[clean_name] = 1
        final_name = clean_name
    sheet_names.append((sheet, final_name))  # 格式：(原始名称, 最终名称)

# 创建Excel文件并分Sheet保存
with pd.ExcelWriter("分类结果.xlsx", engine="openpyxl") as writer:
    for original_name, final_name in sheet_names:
        # 筛选对应分类的数据，并移除"sheet表名"列
        df_subset = filtered_df[filtered_df["sheet表名"] == original_name].drop(columns=["sheet表名"])
        # 写入Sheet
        ordered_columns = ["供料器类型", "合并结果", "物料编码", "物料描述", "用量", "位号", "间距"]
        result_df = df_subset[ordered_columns]
        result_df.columns = ["Feeder类型", "位置", "物料编码", "物料描述", "用量", "位号", "间距"]
        sorted_df = result_df.sort_values(by=['位置',"用量"])
        sorted_df.to_excel(writer, sheet_name=final_name, index=False)







def merge_feeder_by_position(input_path, output_path):
    # 加载工作簿
    wb = load_workbook(input_path)
    
    for sheet in wb:
        # 查找列位置
        position_col = None
        feeder_col = None
        
        # 遍历标题行（假设第一行为标题）
        for cell in sheet[1]:
            if cell.value == "位置":
                position_col = get_column_letter(cell.column)
            elif cell.value == "Feeder类型":
                feeder_col = get_column_letter(cell.column)
        
        # 检查必要列是否存在
        if not position_col or not feeder_col:
            print(f"跳过工作表 {sheet.title}（缺少必要列）")
            continue
        
        # 合并逻辑
        current_group = None
        merge_start = 2  # 数据起始行
        
        for row in range(2, sheet.max_row + 2):  # +2确保处理最后一行
            # 获取位置值并处理空值
            pos_cell = sheet[f"{position_col}{row}"] if row <= sheet.max_row else None
            pos_value = pos_cell.value if pos_cell else None
            
            # 生成分组key
            if pos_value and '-' in pos_value:
                parts = pos_value.split('-')
                group_key = '-'.join(parts[:3]) if len(parts)>=3 else pos_value
            else:
                group_key = pos_value
            
            # 分组变化时执行合并
            if group_key != current_group:
                if current_group is not None and (row-1) > merge_start:
                    sheet.merge_cells(
                        start_row=merge_start,
                        end_row=row-1,
                        start_column=sheet[f"{feeder_col}1"].column,
                        end_column=sheet[f"{feeder_col}1"].column
                    )
                current_group = group_key
                merge_start = row
    
    # 保存结果
    wb.save(output_path)
    # print(f"处理完成，已保存到 {output_path}")


def merge_adjacent_cells(file_path, output_path):
    # 加载工作簿
    wb = load_workbook(file_path)
    
    # 定义需要处理的列标题
    target_columns = ['位置', '用量', '位号', '间距']
    
    for sheet in wb:
        # 创建列位置字典 {列标题: 列字母}
        col_map = {}
        for cell in sheet[1]:  # 假设标题在第一行
            if cell.value in target_columns:
                col_map[cell.value] = get_column_letter(cell.column)
        
        # 如果没有找到关键列则跳过该sheet
        if '位置' not in col_map:
            continue
        
        # 获取所有需要合并的列字母
        merge_cols = [col_map[col] for col in target_columns if col in col_map]
        
        # 合并逻辑
        start_row = 2  # 数据从第二行开始
        prev_value = None
        merge_start = start_row
        
        for row in range(start_row, sheet.max_row + 2):  # +2确保处理最后一行
            current_value = sheet[f"{col_map['位置']}{row}"].value if row <= sheet.max_row else None
            
            if current_value == prev_value:
                continue
                
            if prev_value is not None and (row - 1) > merge_start:
                # 执行合并
                for col in merge_cols:
                    sheet.merge_cells(
                        start_row=merge_start,
                        start_column=sheet[col + str(merge_start)].column,
                        end_row=row - 1,
                        end_column=sheet[col + str(merge_start)].column
                    )
                    
            prev_value = current_value
            merge_start = row
    
    # 保存文件
    wb.save(output_path)

# 使用示例
merge_feeder_by_position("分类结果.xlsx", "分类结果1.xlsx")
merge_adjacent_cells('分类结果1.xlsx', '最终结果.xlsx')



