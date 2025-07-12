import xlrd
import pandas as pd
import numpy as np
import glob
import os
import re
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Alignment, PatternFill, Side
from openpyxl.utils import get_column_letter
from copy import copy


def process_merged_cells_xls(file_path, sheet_name):
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_name(sheet_name)
    
    merged_regions = []
    for merged_range in sheet.merged_cells:
        min_row, max_row, min_col, max_col = merged_range
        value = sheet.cell(min_row, min_col).value
        merged_regions.append((min_row, max_row, min_col, max_col, value))
    
    data = []
    for row_idx in range(sheet.nrows):
        row_data = []
        for col_idx in range(sheet.ncols):
            cell_value = sheet.cell(row_idx, col_idx).value
            for (m_min_row, m_max_row, m_min_col, m_max_col, m_value) in merged_regions:
                if (m_min_row <= row_idx < m_max_row) and (m_min_col <= col_idx < m_max_col):
                    cell_value = m_value
                    break
            row_data.append(cell_value)
        data.append(row_data)
    return data

def extract_data_blocks(data):
    blocks = []
    current_block = None
    direction_search_range = 3  # 在标题行后最多搜索3行寻找方向

    for i, row in enumerate(data):
        # 匹配包含 Line 的标题行
        line_cells = [cell for cell in row if isinstance(cell, str) and 'Line' in cell]

        if line_cells:
            # 保存前一个区块
            if current_block:
                blocks.append(current_block)
            
            # 初始化新区块
            title_cell = line_cells[0]
            current_block = {
                'title': title_cell,
                'direction': None,  # 初始化为空
                'headers': [],
                'rows': [],
                'direction_found': False  # 方向是否已找到的标记
            }

            # 在后续行中搜索方向信息（最多检查3行）
            for j in range(i+1, min(i+1 + direction_search_range, len(data))):
                direction_row = data[j]
                # 匹配独立的方向单元格（纯"左"或"右"）
                direction_cells = [
                    cell for cell in direction_row 
                    if isinstance(cell, str) and re.fullmatch(r'\s*[左右]\s*', str(cell))
                ]
                if direction_cells:
                    current_block['direction'] = direction_cells[0].strip()
                    current_block['direction_found'] = True
                    break  # 找到后停止搜索

        elif current_block:
            # 优先处理方向未找到的情况
            if not current_block['direction_found']:
                # 尝试在本行匹配复合方向信息（如"方向：左"）
                for cell in row:
                    if isinstance(cell, str):
                        match = re.search(r'([左右])', cell)
                        if match:
                            current_block['direction'] = match.group(1)
                            current_block['direction_found'] = True
                            break

            # 列头识别逻辑（仅在方向处理后执行）
            if not current_block['headers']:
                if any('供料器' in str(cell) for cell in row):
                    current_block['headers'] = [str(cell).strip() for cell in row]
                    continue  # 跳过本行后续处理

            # 数据行收集（排除空行）
            if any(cell is not None for cell in row):
                current_block['rows'].append(row)

    # 添加最后一个区块
    if current_block:
        blocks.append(current_block)
    
    return blocks

def daochu():
# 主流程
    bom_folder = os.path.join(os.getcwd(), '导出文件')
    files = glob.glob(os.path.join(bom_folder, '*.xls'))
    if not files:
        print("未找到 .xls 文件，请检查路径！")
        exit()

    file_path = files[0]
    data = process_merged_cells_xls(file_path, 'Sheet1')

    # 调试数据
    # print("==== 前10行数据 ====")
    # for idx, row in enumerate(data[:10]):
    #     print(f"行 {idx}: {row}")

    data_blocks = extract_data_blocks(data)
    print("提取到的数据块数量:", len(data_blocks))

    # 转换为 DataFrame
    dfs = []
    for block in data_blocks:
        if block.get('headers') and block.get('rows'):
            df = pd.DataFrame(block['rows'], columns=block['headers'])
            df['料台标题'] = block['title']
            df['料台左右'] = block['direction']
            dfs.append(df)
        # else:
        #     print("跳过无效块:", block)

    if dfs:
        final_df = pd.concat(dfs, ignore_index=True)
        
        # --------------------------
        # 新增输出结果处理逻辑
        # --------------------------
        # 1. B列（供料器列）向下填充空值
        if '供料器' in final_df.columns:
            final_df['供料器'] = final_df['供料器'].fillna(method='ffill')
        else:
            print("警告: 未找到B列（供料器列）")
        
        target_column = '元件'  # 根据实际列名修改

        # 2. 空值预处理（处理多种空值表现形式）
        null_values = ['', 'NA', 'NaN', 'NULL', 'None', ' ']
        final_df[target_column] = final_df[target_column].replace(null_values, pd.NA)

        # 3. 统计空值情况（调试用）
        # print(f"删除前数据量: {len(final_df)} 行")
        # print(f"'{target_column}'列空值数量: {final_df[target_column].isna().sum()}")

        # 4. 执行删除操作（保留有数据的行）
        if target_column in final_df.columns:
            # 方法一：直接删除空值行（严格模式）
            final_df.dropna(
                subset=[target_column],
                how='any',
                inplace=True
            )
            
            
            # print(f"删除后数据量: {len(final_df)} 行")
            
            # 二次验证
            if final_df[target_column].isna().sum() > 0:
                print("警告: 仍存在空值，请检查数据源")
        else:
            print(f"关键列 '{target_column}' 不存在，可用列：{list(final_df.columns)}")

        # 自动填充空白列名（优化版）
        new_columns = []
        counter = 0
        for col in final_df.columns:
            # 处理空值或默认生成的 Unnamed 列
            if pd.isna(col) or str(col).strip() in ('', 'Unnamed: 0'):
                new_col = f"column_{counter}"
                counter += 1
            else:
                new_col = col
            
            # 统一中文列名（示例：包含 "Line" 的列重命名为 "料台标题"）
            if 'Line' in str(new_col):
                new_col = '料台标题'
            
            new_columns.append(new_col)

        final_df.columns = new_columns

        # 验证列名是否存在（若缺失则动态创建空列）
        required_columns = ["料台标题", "料台左右", "料槽", "供料器类型", "column_8", "元件", "间距"]
        for col in required_columns:
            if col not in final_df.columns:
                final_df[col] = None  # 创建空列或填充默认值

        # 正确选择多列（使用列表语法）
        end_df = final_df[required_columns]

        end_df.replace(['', 'NA', 'NaN', 'null', 'None', ' '], np.nan, inplace=True)

        # 检查数据类型并转换
        for col in end_df.columns:
            if end_df[col].dtype == 'category':
                end_df[col] = end_df[col].astype('object')

        # 重置索引
        end_df = end_df.reset_index(drop=True)

        # 填充NaN，先ffill再bfill处理首行
        filleed_df = end_df.fillna(method='ffill').fillna(method='bfill')

        def merge_with_flag(row):
            name = row['料台标题']
            cao = row['料槽']
            price = row['column_8']
            return f"{name}-{int(cao)}-{int(price)}"

        filleed_df['合并结果'] = filleed_df.apply(merge_with_flag, axis=1)
        
        def merge_flag(row):
            name = row['料台标题']
            left_right = row['料台左右']

            if left_right == '左':
                return f"{name}-2"
            elif left_right == '右':
                return f"{name}-1"


        filleed_df['sheet表名'] = filleed_df.apply(merge_flag, axis=1)
        
        # 保存结果
        # filleed_df.to_csv("output.csv", index=False, encoding='utf-8-sig')
        # print("处理完成，已保存到 output.csv")
        return filleed_df
    else:
        print("警告: 无有效数据块可合并！")













def bom():

    logging.basicConfig(level=logging.INFO)

    # 配置路径和关键词
    bom_folder = os.path.join(os.getcwd(), 'BOM')
    files = glob.glob(os.path.join(bom_folder, '*[Bb][Oo][Mm]*.xlsx'))
    required_columns = {'编码', '描述', '位号'}
    keyword_map = {'编码': '物料编码', '描述': '物料描述', '位号': '位号', '用量': '用量'}

    data_frames = []

    for file in files:
        try:
            xls = pd.ExcelFile(file)
        except Exception as e:
            logging.error(f"读取文件失败: {file}, 错误: {e}")
            continue
        
        for sheet_name in xls.sheet_names:
            try:
                # 动态寻找表头行（前10行内）
                header_row = None
                for skiprows in range(0, 10):
                    df_temp = pd.read_excel(xls, sheet_name, skiprows=skiprows, nrows=1)
                    row_columns = df_temp.columns.str.lower().tolist()
                    # 检查是否包含所有必需列的关键词
                    found = all(any(kw in col for col in row_columns) for kw in required_columns)
                    if found:
                        header_row = skiprows
                        break
                
                if header_row is None:
                    logging.info(f"跳过 {file} 的 {sheet_name}: 未找到表头")
                    continue
                
                # 读取数据
                df = pd.read_excel(xls, sheet_name, skiprows=header_row)
                if df.empty:
                    logging.info(f"文件 {file} 的 {sheet_name} 无数据")
                    continue
                
                # 添加来源信息
                df['来源文件'] = os.path.basename(file)
                df['来源Sheet'] = sheet_name
                # --- 修改后的合并位号列代码 ---
                ref_cols = [col for col in df.columns if '位号' in col]  # 改为包含"位号"的列名
                if len(ref_cols) > 0:
                    def merge_ref_des(row):
                        refs = []
                        for col in ref_cols:
                            val = row[col]
                            if pd.notna(val):
                                # 处理可能的换行符和多个空格
                                parts = str(val).replace('\n', ' ').split()  # 自动处理连续空格和换行
                                refs.extend([p.strip() for p in parts if p.strip()])
                        # 按原始顺序去重（替代set的无序性）
                        seen = set()
                        ordered_refs = []
                        for ref in refs:
                            if ref not in seen:
                                seen.add(ref)
                                ordered_refs.append(ref)
                        return ' '.join(ordered_refs) if ordered_refs else ''
                    
                    df['位号'] = df.apply(merge_ref_des, axis=1)
                    df.drop(ref_cols, axis=1, inplace=True)
                data_frames.append(df)
                    
                # --- 列名重命名（确保唯一性）---
                sorted_keywords = sorted(keyword_map.keys(), key=lambda x: -len(x))
                new_columns = []
                seen = set()
                for col in df.columns:
                    col_str = str(col)
                    col_lower = col_str.lower()
                    matched = False
                    for kw in sorted_keywords:
                        if kw in col_lower:
                            base_name = keyword_map[kw]
                            if base_name in seen:
                                suffix = 1
                                while f"{base_name}.{suffix}" in seen:
                                    suffix += 1
                                new_name = f"{base_name}.{suffix}"
                            else:
                                new_name = base_name
                            new_columns.append(new_name)
                            seen.add(new_name)
                            matched = True
                            break
                    if not matched:
                        new_columns.append(col_str)
                        seen.add(col_str)
                df.columns = new_columns
                

            except Exception as e:
                logging.warning(f"处理 {file} 的 {sheet_name} 出错: {e}")

    # 合并和清洗数据
    if data_frames:
        combined_df = pd.concat(data_frames, ignore_index=True)
        combined_df = combined_df.dropna(axis=1, how='all')
        final_columns = list(keyword_map.values()) + ['来源文件', '来源Sheet']
        final_df = combined_df.reindex(columns=final_columns, fill_value='')
        
        final_df['位号'] = (
        final_df['位号']
        .astype(str)
        .replace(['nan', '', ' '], pd.NA)  # 统一替换为 Pandas 空值
    )
        final_df['位号'] = final_df['位号'].fillna(method='ffill')
        print("合并后的数据样例：")
        print(final_df.head())
        # final_df.to_excel("excel.xlsx")
        return final_df
    else:
        print("未找到有效数据")


daochu_c = daochu()
bom_c = bom()
bom_c['物料编码'] = bom_c['物料编码'].astype(str).str.strip()
daochu_c['元件'] = daochu_c['元件'].astype(str).str.strip()
common = set(bom_c['物料编码']) & set(daochu_c['元件'])
# 筛选 common 对应的位号
common_bom = bom_c[bom_c['物料编码'].isin(common)]
common_ref_des = common_bom['位号'].unique()

# 清洗位号并生成集合
common_ref_set = {
    str(ref).strip()
    for ref in common_ref_des
    if pd.notna(ref) and str(ref).strip() != ''
}

# 处理位号列（统一分隔符为逗号）
bom_c['位号'] = (
    bom_c['位号']
    .fillna('')
    .astype(str)
    .str.strip()
    .str.replace(r'[;\t]+', ',', regex=True)
)

# 按逗号拆分位号
bom_c['位号_split'] = bom_c['位号'].str.split(',')

# 筛选匹配的行
result = bom_c[
    bom_c['位号_split'].apply(
        lambda x: bool(set(map(str.strip, x)) & common_ref_set)
    )
].drop(columns=['位号_split'])

# 数据预处理：统一列格式
result['物料编码'] = result['物料编码'].astype(str).str.strip()
daochu_c['元件'] = daochu_c['元件'].astype(str).str.strip()

# 执行左连接（保留 result 所有行）
merged_df = pd.merge(
    result,
    daochu_c,
    how='left',          # 左连接
    left_on='物料编码',   # result 表的连接键
    right_on='元件',      # daochu_c 表的连接键
    suffixes=('', '_右表')  # 重命名重复列
)

merged_df['sheet表名'] = (
        merged_df['sheet表名']
        .astype(str)
        .replace(['nan', '', ' '], pd.NA)  # 统一替换为 Pandas 空值
    )
merged_df['sheet表名'] = merged_df['sheet表名'].fillna(method='ffill')






# ---------------------- 初始化配置 ----------------------
template_path = 'template.xlsx'
output_path = 'output.xlsx'

# ---------------------- 准备模板数据 ----------------------
# 加载模板文件
template_wb = load_workbook(template_path)
template_ws = template_wb.active

# 存储模板前6行格式
header_rows = []
for row in template_ws.iter_rows(min_row=1, max_row=6):
    row_data = []
    for cell in row:
        row_data.append({
            'value': cell.value,
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'alignment': copy(cell.alignment)
        })
    header_rows.append(row_data)

# 获取模板样式（假设数据行在模板第7、8行）
header_style = {
    'font': copy(template_ws['A7'].font),
    'border': copy(template_ws['A7'].border),
    'fill': copy(template_ws['A7'].fill),
    'alignment': copy(template_ws['A7'].alignment)
}

data_style = {
    'font': copy(template_ws['A8'].font),
    'border': copy(template_ws['A8'].border),
    'fill': copy(template_ws['A8'].fill),
    'alignment': copy(template_ws['A8'].alignment)
}

# ---------------------- 主处理逻辑 ----------------------
def process_data(writer):
    # 列映射关系
    column_mapping = {
        '供料器类型': 'Feeder类型',
        '合并结果': '位置',
        '物料编码': '物料编码',
        '物料描述': '物料描述', 
        '用量': '用量',
        '位号': '位号',
        '间距': '间距'
    }

    # 分组处理
    grouped = merged_df.groupby('sheet表名')
    
    for sheet_name, group in grouped:
        # 清理sheet名称
        safe_sheet_name = str(sheet_name).replace('/', '-')[:31]
        
        # 创建新sheet
        if safe_sheet_name in writer.book.sheetnames:
            del writer.book[safe_sheet_name]
        ws = writer.book.create_sheet(safe_sheet_name)

        # ---------------------- 插入模板行 ----------------------
        # 插入前6行模板内容
        for row_idx, row_data in enumerate(header_rows, 1):
            for col_idx, cell_data in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_data['value'])
                cell.font = cell_data['font']
                cell.border = cell_data['border']
                cell.fill = cell_data['fill']
                cell.alignment = cell_data['alignment']

        # ---------------------- 准备数据 ----------------------
        final_df = (
            group[column_mapping.keys()]
            .rename(columns=column_mapping)
            [['Feeder类型', '位置', '物料编码', '物料描述', '用量', '位号', '间距']]
            .fillna('')
            .reset_index(drop=True)
        )

        # ---------------------- 写入标题行 ----------------------
        for col_idx, col_name in enumerate(final_df.columns, 1):
            cell = ws.cell(row=7, column=col_idx, value=col_name)
            cell.font = header_style['font']
            cell.border = header_style['border']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']

        # ---------------------- 写入数据行 ----------------------
        start_data_row = 8
        for df_idx, (_, row) in enumerate(final_df.iterrows()):
            excel_row = start_data_row + df_idx
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font = data_style['font']
                cell.border = data_style['border']
                cell.fill = data_style['fill']
                cell.alignment = data_style['alignment']

        # ---------------------- 合并单元格逻辑 ----------------------
        thin_border = Side(border_style="thin", color="000000")
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        feeder_col = final_df.columns.get_loc('Feeder类型') + 1  # openpyxl列从1开始
        total_data_rows = final_df.shape[0]
        current_excel_row = start_data_row

        while current_excel_row < start_data_row + total_data_rows:
            # 查找合并起始点
            start_merge_row = current_excel_row
            while current_excel_row < start_data_row + total_data_rows:
                cell_value = ws.cell(row=current_excel_row, column=feeder_col).value
                if cell_value not in [None, '']:
                    break
                current_excel_row += 1
            else:
                break

            # 查找合并结束点
            end_merge_row = current_excel_row + 1
            while end_merge_row <= start_data_row + total_data_rows:
                cell_value = ws.cell(row=end_merge_row, column=feeder_col).value
                if cell_value not in [None, '']:
                    break
                end_merge_row += 1

            # 执行合并
            if end_merge_row > start_merge_row + 1:
                # 合并Feeder类型列
                ws.merge_cells(
                    start_row=start_merge_row,
                    end_row=end_merge_row - 1,
                    start_column=feeder_col,
                    end_column=feeder_col
                )
                ws.cell(start_merge_row, feeder_col).alignment = center_alignment

                # 合并其他列
                for col in ['位置', '用量', '位号', '间距']:
                    col_idx = final_df.columns.get_loc(col) + 1
                    ws.merge_cells(
                        start_row=start_merge_row,
                        end_row=end_merge_row - 1,
                        start_column=col_idx,
                        end_column=col_idx
                    )
                    ws.cell(start_merge_row, col_idx).alignment = center_alignment

            current_excel_row = end_merge_row

        # ---------------------- 调整列宽 ----------------------
        for col_idx, col_name in enumerate(final_df.columns, 1):
            max_len = final_df[col_name].astype(str).apply(len).max()
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len, len(col_name)) + 2


    # 初始化写入器
writer = pd.ExcelWriter(output_path, engine='openpyxl')
writer.book = load_workbook(template_path)

# 处理数据
process_data(writer)

# 清理默认sheet
if 'Sheet' in writer.book.sheetnames:
    del writer.book['Sheet']

# 保存结果
writer.close()