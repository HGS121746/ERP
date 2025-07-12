import pandas as pd
import re,os
import openpyxl
import glob
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

# 读取文件（注意文件路径和编码）
bom_folder = os.path.join(os.getcwd(), '上料表')  # 获取 '模板文件' 文件夹的绝对路径
files = glob.glob(
    os.path.join(bom_folder, '*.csv')  # 匹配模式：任意csv文件
)
file_path = files[0]
with open(file_path, 'r', encoding='gbk') as f:
    lines = f.readlines()

# 预处理：跳过文件开头的元数据行（前5行）
lines = lines[4:]

blocks = []
current_section = None
current_block = []
header = None

for line in lines:
    line = [cell.strip() for cell in line.strip().split(',')]  # 清理空格
    # 检测区块起始行（如 [#1(RX-8)]）
    section_cells = [cell for cell in line if re.search(r'#\d+', cell)]
    if section_cells:
        if current_block and header is not None:
            block_df = pd.DataFrame(current_block, columns=header)
            block_df['区块号'] = current_section
            blocks.append(block_df)
            current_block = []
        # 提取区块号
        section_cell = section_cells[0]
        current_section = re.search(r'#(\d+)', section_cell).group(1)
        header = None
    # 检测表头行（包含 "No" 和 "供应" 等中文字段）
    elif "No" in line and "供应" in line and "编号" in line:
        # 标准化表头（中日文映射）
        header = ["No", "供应", "编号", "元件名", "供料器类型", "贴片数", "贴片ID"]
    # 数据行处理
    elif len(line) >= 7 and line[0].isdigit():
        if header is not None:
            current_block.append(line[:7])  # 取前7列有效数据

# 添加最后一个区块
if current_block and header is not None:
    block_df = pd.DataFrame(current_block, columns=header)
    block_df['区块号'] = current_section
    blocks.append(block_df)

# 合并所有区块数据
combined_df = pd.concat(blocks, ignore_index=True)

# Feeder列处理
def extract_feeder(s):
    parts = str(s).split()
    return parts[1] if (parts[0] == 'RF' and len(parts) > 1) else parts[0]

combined_df['Feeder'] = combined_df['供料器类型'].apply(extract_feeder)

# 位置列处理
supply_mapping = {
    '左前面': '2', '右前面': '1'
}
# combined_df['区块号'] = combined_df['区块号'].astype(str)


# 定义位置生成规则（新增区块号判断）
combined_df['位置'] = combined_df.apply(
    lambda x: (
        # 当区块号 >2 时：区块号-编号
        f"{x['区块号']}-{x['编号']}"
        # 当区块号 <=2 时：区块号-供应映射-编号
        if int(x['区块号']) > 2 
        else f"{x['区块号']}-{supply_mapping.get(x['供应'], '0')}-{x['编号']}"
    ),
    axis=1
)



# 间距列处理
def extract_spacing(s):
    try:
        # 提取括号内容
        match = re.search(r'\((.*?)\)', str(s))
        if not match:
            return ''
        
        # 计算表达式
        expression = match.group(1).strip()
        return eval(expression)  # 返回计算结果
    except:
        # 表达式无效或计算异常时返回空
        return ''

combined_df['间距'] = combined_df['供料器类型'].apply(extract_spacing)

# 生成最终表格
final_df = combined_df[['Feeder', '位置','元件名', '间距']]

# 输出结果
# print(final_df.head())
final_df.to_excel('清洗后的上料表.xlsx', index=False)






# 获取当前目录下所有包含"bom"的Excel文件（不区分大小写）
bom_folder = os.path.join(os.getcwd(), 'BOM')  # 获取 'BOM' 文件夹的绝对路径
files = glob.glob(
    os.path.join(bom_folder, '*.xlsx')  # 匹配模式：文件名含 BOM（忽略大小写）
)
# files = glob.glob('*[Bb][Oo][Mm]*.xlsx')

# 定义必须存在的列名关键词集合
required_columns = {'编码', '描述', '位号'}

# 初始化数据存储列表
data_frames = []

# 遍历每个Excel文件
for file in files:
    # 使用openpyxl加载工作簿
    wb = load_workbook(file, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 定位包含所有必要关键词的表头行
        header_row = None
        for row_idx in range(1, ws.max_row + 1):
            row_values = [cell.value for cell in ws[row_idx]]
            if all(col.lower() in str(row_values).lower() for col in required_columns):
                header_row = row_idx
                break
        
        if not header_row:
            continue  # 跳过没有找到表头的sheet
        
        # 提取表头和数据
        headers = [cell.value for cell in ws[header_row]]
        data_rows = ws.iter_rows(min_row=header_row + 1, values_only=True)
        
        # 创建DataFrame并重置索引
        df = pd.DataFrame(data_rows, columns=headers).reset_index(drop=True)
        
        # 添加来源信息
        df['来源文件'] = file
        df['来源Sheet'] = sheet_name
        
        # 确保索引唯一性
        df = df.loc[:, ~df.columns.duplicated()]  # 删除重复列
        df.reset_index(drop=True, inplace=True)
        data_frames.append(df)

# 合并所有数据
if data_frames:  # 确保列表不为空
    combined_df = pd.concat(data_frames, ignore_index=True)
    # print(combined_df.head())
    # 数据清洗
    selected_columns = combined_df.filter(regex='编码|描述|位号|用量').copy()
    

    # 筛选列名包含指定关键词的列
    keyword_map = {
    '编码': '物料编码',
    '描述': '物料描述',
    '位号': '位号',
    '用量': '用量'
    }

    # 重命名列
    renamed_columns = []
    for col in selected_columns.columns:
        for keyword, new_name in keyword_map.items():
            if keyword in col:
                renamed_columns.append(new_name)
                break
        else:
            renamed_columns.append(col)  # 若未匹配到关键词，保留原名（理论上不会执行）

    # 更新 DataFrame 的列名
    selected_columns.columns = renamed_columns
    selected_columns.to_excel('清洗后的bom表.xlsx', index=False)
    # print(selected_columns.head())

else:
    print("未找到符合条件的表格数据")

# merged_df = pd.merge(
#     final_df,
#     selected_columns,
#     left_on='元件名',       # 左表的连接键
#     right_on='物料编码',    # 右表的连接键
#     how='left'             # 左连接模式
# )
# 处理右表：去除物料编码的小数点
selected_columns['物料编码'] = selected_columns['物料编码'].astype(str).str.replace('.', '', regex=False)

# 确保左表元件名为字符串类型
final_df['元件名'] = final_df['元件名'].astype(str)

# 逐行处理合并
merged_rows = []
for _, left_row in final_df.iterrows():
    component = left_row['元件名']
    mask = selected_columns['物料编码'].str.contains(component, na=False, regex=False)
    matched = selected_columns[mask]
    if not matched.empty:
        # 取第一个匹配项（可选：合并所有匹配项）
        merged_row = {**left_row.to_dict(), **matched.iloc[0].to_dict()}
    else:
        merged_row = {**left_row.to_dict(), **{col: None for col in selected_columns.columns}}
    merged_rows.append(merged_row)

merged_df = pd.DataFrame(merged_rows)
# 查看合并后的结果（可选）
# merged_df.to_excel('merged_df.xlsx')
# print(merged_df.head())




# 读取模板文件
bom_folder = os.path.join(os.getcwd(), '模板文件')  # 获取 '模板文件' 文件夹的绝对路径
files = glob.glob(
    os.path.join(bom_folder, '*.xlsx')  # 匹配模式：任意xlsx文件
)
template_path = files[0]
wb = openpyxl.load_workbook(template_path)

# 读取final_df数据（假设已生成）
# final_df = ...

# 按"位置"列的第一个数字拆分数据
sheets_data = {}
for idx, row in merged_df.iterrows():
    position = row['位置']
    sheet_num = position.split('-')[0]  # 提取第一个数字作为工作表编号
    sheet_key = str(sheet_num)
    
    if sheet_key not in sheets_data:
        sheets_data[sheet_key] = []
    sheets_data[sheet_key].append(row)

# 将数据写入对应工作表
for sheet_num, data in sheets_data.items():
    # 转换为DataFrame
    df = pd.DataFrame(data).reset_index(drop=True)
    df = df[['Feeder', '位置', '元件名', '间距','物料描述','用量','位号']]
    
    # 添加缺失列（留空）
    df['新物料编码'] = df['元件名']
    # df['物料描述'] = df['物料描述']
    # df['用量'] = ''
    # df['位号'] = ''
    
    # 按模板列顺序调整
    df = df[['Feeder', '位置', '新物料编码', '物料描述', '用量', '位号', '间距']]
    
    # 获取对应工作表（sheet_num对应metadata.sheet_index_num）
    sheet_name = wb.sheetnames[int(sheet_num)-1]  # 假设sheet顺序与编号一致
    ws = wb[sheet_name]
    
    # 从第9行开始写入数据（跳过前8行表头）
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=9):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

# 保存为新文件（不覆盖模板）
output_path = 'result.xlsx'
wb.save(output_path)

